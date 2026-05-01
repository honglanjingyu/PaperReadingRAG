# app/service/core/deepdoc/parser/excel_parser.py
from io import BytesIO
from openpyxl import load_workbook

class RAGFlowExcelParser:
    """Excel 文档解析器 (XLSX/XLS)"""

    def __call__(self, fnm):
        """解析 Excel 文件"""
        if isinstance(fnm, str):
            wb = load_workbook(fnm)
        else:
            # 验证并解析二进制数据
            if not fnm.startswith(b'PK'):
                raise ValueError("不是有效的 XLSX 文件格式")
            wb = load_workbook(BytesIO(fnm))

        res = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue
            ti = list(rows[0])  # 表头
            for r in list(rows[1:]):  # 数据行
                fields = []
                for i, c in enumerate(r):
                    if not c.value:
                        continue
                    t = str(ti[i].value) if i < len(ti) else ""
                    t += ("：" if t else "") + str(c.value)
                    fields.append(t)
                line = "; ".join(fields)
                res.append(line)
        return res

    def html(self, fnm, chunk_rows=256):
        """将 Excel 转换为 HTML 表格格式"""
        # 用于大表格的分块处理
        pass